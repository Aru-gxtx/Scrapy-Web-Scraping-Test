import argparse
import json
import re
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook


def normalize_catalog_number(value):
    if value is None:
        return ""

    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""

    text = unquote(text)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^a-z0-9 ]", "", text)
    return text


def get_first_from_list(value):
    if isinstance(value, list):
        if value:
            return str(value[0]).strip()
        return ""
    if value:
        return str(value).strip()
    return ""


def build_payload(item):
    attributes = item.get("attributes") or {}
    description = item.get("description") or ""
    downloads = item.get("downloads") or []
    image_links = item.get("image_links") or []
    
    # Get first image link if available (from API data)
    image_link = ""
    if image_links and isinstance(image_links, list) and len(image_links) > 0:
        image_link = image_links[0]
    
    # Get first download URL if available
    download_url = ""
    if downloads and isinstance(downloads, list) and len(downloads) > 0:
        download_url = downloads[0].get("url", "")
        # Convert relative URL to absolute URL if needed
        if download_url and download_url.startswith("/"):
            download_url = f"https://www.steelite-utopia.com{download_url}"
    
    # Extract attributes (they're stored as lists in API)
    features = ", ".join(attributes.get("features", []))
    colour = get_first_from_list(attributes.get("colour"))
    material = get_first_from_list(attributes.get("material"))
    barcode = get_first_from_list(attributes.get("barcodeOuter"))
    pattern = get_first_from_list(attributes.get("range"))
    brand = get_first_from_list(attributes.get("brand"))
    
    payload = {
        "Image Link": image_link,
        "Overview": description,
        "Length": "",  # Not available in API
        "Width": "",   # Not available in API
        "Height": "",  # Not available in API
        "Capacity": "", # Not available in API
        "Features": features,
        "Edge Style": "",  # Not available in API
        "Volume": "",  # Not available in API
        "Diameter": "", # Not available in API
        "Color": colour,
        "Shape": "",  # Not available in API
        "Material": material,
        "Pattern": pattern,
        "EAN Code": "",  # Not available in API
        "Barcode": barcode,
        "Brand": brand,  # Additional field from API
    }
    
    # Add download link if column exists
    if download_url:
        payload["Download Link"] = download_url
    
    return payload


def load_json_entries(json_path):
    with open(json_path, "r", encoding="utf-8") as file:
        return json.load(file)


def populate_workbook(excel_path, json_entries):
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    headers = {
        (sheet.cell(row=1, column=column).value or "").strip(): column
        for column in range(1, sheet.max_column + 1)
    }

    catalog_column = headers.get("Mfr Catalog No.")
    if not catalog_column:
        raise ValueError("Column 'Mfr Catalog No.' was not found in the workbook.")

    item_description_column = headers.get("Item Description")
    start_column = 5

    # Build map of normalized catalog -> list of row numbers in Excel
    excel_catalog_map = {}
    # Build fallback list for rows with blank Mfr Catalog No. (last-resort matching)
    blank_catalog_rows = []

    for row in range(2, sheet.max_row + 1):
        raw_catalog = sheet.cell(row=row, column=catalog_column).value
        catalog_number = normalize_catalog_number(raw_catalog)

        if catalog_number:
            if catalog_number not in excel_catalog_map:
                excel_catalog_map[catalog_number] = []
            excel_catalog_map[catalog_number].append(row)
            continue

        if item_description_column:
            raw_description = sheet.cell(row=row, column=item_description_column).value
            normalized_description = normalize_text(raw_description)
            if normalized_description:
                blank_catalog_rows.append((row, normalized_description))

    matched_entries = 0
    updated_cells = 0
    used_excel_rows = set()

    for json_item in json_entries:
        json_catalog = normalize_catalog_number(json_item.get("catalog_number"))
        target_row = None

        if json_catalog:
            matching_excel_rows = excel_catalog_map.get(json_catalog, [])

            # Find first Excel row with this catalog that hasn't been used yet
            for row in matching_excel_rows:
                if row not in used_excel_rows:
                    target_row = row
                    used_excel_rows.add(row)
                    break

            if not target_row and matching_excel_rows:
                # All rows with this catalog have been used; reuse the first one
                target_row = matching_excel_rows[0]
        else:
            # Last resort: only for JSON entries without Mfr Catalog No.
            product_name = normalize_text(json_item.get("product_name"))
            if product_name and blank_catalog_rows:
                best_row = None
                best_score = 0.0

                for row, description in blank_catalog_rows:
                    if row in used_excel_rows:
                        continue

                    score = SequenceMatcher(None, product_name, description).ratio()
                    if product_name in description or description in product_name:
                        score += 0.2

                    if score > best_score:
                        best_score = score
                        best_row = row

                if best_row and best_score >= 0.55:
                    target_row = best_row
                    used_excel_rows.add(best_row)

        if not target_row:
            continue

        matched_entries += 1
        payload = build_payload(json_item)

        for header_name, new_value in payload.items():
            if not new_value:
                continue

            target_column = headers.get(header_name)
            if not target_column or target_column < start_column:
                continue

            cell = sheet.cell(row=target_row, column=target_column)
            # Overwrite existing data with new API values
            old_value = cell.value
            cell.value = new_value
            if old_value != new_value:
                updated_cells += 1

    workbook.save(excel_path)
    return matched_entries, updated_cells


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Populate empty cells in STEELITE.xlsx using Steelite-Utopia API data, "
            "matched by 'Mfr Catalog No.' and limited to columns E onward."
        )
    )
    parser.add_argument(
        "--excel",
        default="sources/STEELITE.xlsx",
        help="Path to STEELITE.xlsx",
    )
    parser.add_argument(
        "--json",
        default="steelite/steelite_utopia_from_api.json",
        help="Path to Steelite-Utopia API JSON file",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    json_path = Path(args.json)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if not json_path.exists():
        raise FileNotFoundError(f"JSON file not found: {json_path}")

    print(f"Loading API data from {json_path}...")
    json_entries = load_json_entries(json_path)
    print(f"Loaded {len(json_entries)} entries from API JSON")
    
    print(f"Populating {excel_path}...")
    matched_entries, updated_cells = populate_workbook(excel_path, json_entries)

    print("=" * 60)
    print("RESULTS")
    print("=" * 60)
    print(f"Matched JSON entries: {matched_entries}")
    print(f"Cells updated: {updated_cells}")
    print("Done. Existing data was overwritten with new API values.")


if __name__ == "__main__":
    main()
