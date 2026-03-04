from openpyxl import load_workbook

wb = load_workbook('sources/STEELITE.xlsx', data_only=True)
ws = wb.active
headers = [(ws.cell(1,c).value or '').strip() for c in range(1, ws.max_column+1)]

# Check a few rows that should be populated
img_col = headers.index('Image Link') + 1 if 'Image Link' in headers else None
overview_col = headers.index('Overview') + 1 if 'Overview' in headers else None

print('Sample populated rows:')
count_populated = 0
for r in range(2, min(100, ws.max_row+1)):
    img = ws.cell(r, img_col).value if img_col else None
    overview = ws.cell(r, overview_col).value if overview_col else None
    if img or overview:
        count_populated += 1
        if count_populated <= 3:
            print(f'  Row {r}: Image={bool(img)}, Overview length={len(str(overview)) if overview else 0}')

print(f'\nTotal rows with populated Image Link or Overview in first 100: {count_populated}')
print(f'Workbook max row: {ws.max_row}')
