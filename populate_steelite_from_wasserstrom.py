import argparse
import json
import re
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook


def normalize_catalog_number(value):
    if value is None:
        return ""

    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""

    text = unquote(text)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^a-z0-9 ]", "", text)
    return text


def build_payload(item):
    product_name = (item.get("product_name") or "").strip()
    overview = (item.get("overview") or "").strip()

    if not overview:
        overview = product_name

    # Keep match context visible when fuzzy redirects were used.
    match_note = ""
    if item.get("exact_match") is False:
        searched = (item.get("searched_catalog_number") or "").strip()
        found = (item.get("catalog_number") or "").strip()
        if searched and found:
            match_note = f"Wasserstrom fuzzy match: searched {searched}, found {found}"

    payload = {
        "Image Link": (item.get("image_link") or "").strip(),
        "Overview": overview,
        "Length": (item.get("length") or "").strip(),
        "Width": (item.get("width") or "").strip(),
        "Height": (item.get("height") or "").strip(),
        "Capacity": (item.get("volume") or "").strip(),
        "Features": match_note,
        "Edge Style": "",
        "Volume": (item.get("volume") or "").strip(),
        "Diameter": (item.get("diameter") or "").strip(),
        "Color": (item.get("color") or "").strip(),
        "Shape": "",
        "Material": (item.get("material") or "").strip(),
        "Pattern": (item.get("pattern") or "").strip(),
        "EAN Code": (item.get("ean_code") or "").strip(),
        "Barcode": (item.get("barcode") or "").strip(),
        # Optional columns: these are only written if present in the workbook headers.
        "Country Of Origin": (item.get("country_of_origin") or "").strip(),
        "Warranty": (item.get("warranty") or "").strip(),
        "Manufacturer": (item.get("manufacturer") or "").strip(),
        "Item Number": (item.get("item_number") or "").strip(),
        "Product URL": (item.get("product_url") or "").strip(),
        "Search URL": (item.get("search_url") or "").strip(),
    }

    return payload


def load_json_entries(json_path):
    with open(json_path, "r", encoding="utf-8") as file:
        return json.load(file)


def populate_workbook(excel_path, json_entries):
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    headers = {
        (sheet.cell(row=1, column=column).value or "").strip(): column
        for column in range(1, sheet.max_column + 1)
    }

    catalog_column = headers.get("Mfr Catalog No.")
    if not catalog_column:
        raise ValueError("Column 'Mfr Catalog No.' was not found in the workbook.")

    item_description_column = headers.get("Item Description")
    start_column = 5

    excel_catalog_map = {}
    blank_catalog_rows = []

    for row in range(2, sheet.max_row + 1):
        raw_catalog = sheet.cell(row=row, column=catalog_column).value
        catalog_number = normalize_catalog_number(raw_catalog)

        if catalog_number:
            if catalog_number not in excel_catalog_map:
                excel_catalog_map[catalog_number] = []
            excel_catalog_map[catalog_number].append(row)
            continue

        if item_description_column:
            raw_description = sheet.cell(row=row, column=item_description_column).value
            normalized_description = normalize_text(raw_description)
            if normalized_description:
                blank_catalog_rows.append((row, normalized_description))

    matched_entries = 0
    updated_cells = 0
    used_excel_rows = set()

    for json_item in json_entries:
        if json_item.get("found") is False:
            continue

        json_catalog = normalize_catalog_number(
            json_item.get("catalog_number") or json_item.get("searched_catalog_number")
        )
        target_rows = []

        if json_catalog:
            matching_excel_rows = excel_catalog_map.get(json_catalog, [])
            # Apply this JSON entry to all Excel rows with this catalog number
            target_rows = matching_excel_rows
        else:
            product_name = normalize_text(json_item.get("product_name"))
            if product_name and blank_catalog_rows:
                best_row = None
                best_score = 0.0

                for row, description in blank_catalog_rows:
                    if row in used_excel_rows:
                        continue

                    score = SequenceMatcher(None, product_name, description).ratio()
                    if product_name in description or description in product_name:
                        score += 0.2

                    if score > best_score:
                        best_score = score
                        best_row = row

                if best_row and best_score >= 0.55:
                    target_rows = [best_row]
                    used_excel_rows.add(best_row)

        if not target_rows:
            continue

        matched_entries += 1
        payload = build_payload(json_item)

        for target_row in target_rows:
            for header_name, new_value in payload.items():
                if not new_value:
                    continue

                target_column = headers.get(header_name)
                if not target_column or target_column < start_column:
                    continue

                cell = sheet.cell(row=target_row, column=target_column)
                old_value = cell.value
                cell.value = new_value
                if old_value != new_value:
                    updated_cells += 1

    workbook.save(excel_path)
    return matched_entries, updated_cells


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Populate empty cells in STEELITE.xlsx using wasserstrom scraped data, "
            "matched by 'Mfr Catalog No.' and limited to columns E onward."
        )
    )
    parser.add_argument(
        "--excel",
        default="sources/STEELITE.xlsx",
        help="Path to STEELITE.xlsx",
    )
    parser.add_argument(
        "--json",
        default="steelite/wasserstrom.json",
        help="Path to wasserstrom JSON file",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    json_path = Path(args.json)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if not json_path.exists():
        raise FileNotFoundError(f"JSON file not found: {json_path}")

    json_entries = load_json_entries(json_path)
    matched_entries, updated_cells = populate_workbook(excel_path, json_entries)

    print(f"Matched JSON entries: {matched_entries}")
    print(f"Cells updated: {updated_cells}")
    print("Done. Existing data was overwritten for matched rows.")


if __name__ == "__main__":
    main()
