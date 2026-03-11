import argparse
import json
import re
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook


def normalize_catalog_number(value):
    if value is None:
        return ""

    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""

    text = unquote(text)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^a-z0-9 ]", "", text)
    return text


def is_steelite_item(item):
    name = normalize_text(item.get("product_name") or item.get("Product Title"))
    url = normalize_text(item.get("product_url") or item.get("Product URL"))
    return "steelite" in name or "steelite" in url


def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def build_payload(item):
    product_name = clean_value(item.get("product_name") or item.get("Product Title"))
    overview = clean_value(item.get("Overview") or item.get("overview"))

    if not overview:
        overview = product_name

    payload = {
        "Image Link": clean_value(item.get("Image Link") or item.get("image_link")),
        "Overview": overview,
        "Length": clean_value(item.get("Length") or item.get("length")),
        "Width": clean_value(item.get("Width") or item.get("width")),
        "Height": clean_value(item.get("Height") or item.get("height")),
        "Capacity": clean_value(item.get("Capacity") or item.get("Volume") or item.get("volume")),
        "Features": "",
        "Edge Style": "",
        "Volume": clean_value(item.get("Volume") or item.get("volume")),
        "Diameter": clean_value(item.get("Diameter") or item.get("diameter")),
        "Color": clean_value(item.get("Color") or item.get("color")),
        "Shape": clean_value(item.get("Shape") or item.get("shape")),
        "Material": clean_value(item.get("Material") or item.get("material")),
        "Pattern": clean_value(item.get("Pattern") or item.get("pattern")),
        "EAN Code": clean_value(item.get("EAN Code") or item.get("ean_code")),
        "Barcode": clean_value(item.get("Barcode") or item.get("barcode")),
        # Optional columns: these are only written if present in the workbook headers.
        "Price": clean_value(item.get("price")),
        "Product URL": clean_value(item.get("product_url") or item.get("Product URL")),
        "Search URL": clean_value(item.get("search_url") or item.get("Search URL")),
    }

    return payload


def load_json_entries(json_path):
    with open(json_path, "r", encoding="utf-8") as file:
        return json.load(file)


def load_incomplete_catalog_numbers(path):
    raw = path.read_text(encoding="utf-8")
    # This file may contain NaN tokens which are invalid JSON.
    raw = re.sub(r"\bNaN\b", "null", raw)
    values = json.loads(raw)

    catalogs = []
    for value in values:
        normalized = normalize_catalog_number(value)
        if normalized:
            catalogs.append(normalized)
    return catalogs


def extract_item_catalog(item):
    return normalize_catalog_number(
        item.get("mpn")
        or item.get("catalog_number")
        or item.get("searched_catalog_number")
        or item.get("Barcode")
        or item.get("barcode")
    )


def payload_score(item):
    payload = build_payload(item)
    return sum(1 for value in payload.values() if str(value).strip())


def build_best_item_by_catalog(json_entries):
    best_by_catalog = {}

    for item in json_entries:
        if item.get("found") is False or item.get("blocked") is True:
            continue
        if not is_steelite_item(item):
            continue

        catalog = extract_item_catalog(item)
        if not catalog:
            continue

        current = best_by_catalog.get(catalog)
        if current is None or payload_score(item) > payload_score(current):
            best_by_catalog[catalog] = item

    return best_by_catalog


def populate_workbook(excel_path, json_entries, incomplete_catalog_numbers):
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    headers = {
        (sheet.cell(row=1, column=column).value or "").strip(): column
        for column in range(1, sheet.max_column + 1)
    }

    catalog_column = headers.get("Mfr Catalog No.")
    if not catalog_column:
        raise ValueError("Column 'Mfr Catalog No.' was not found in the workbook.")

    start_column = 5
    incomplete_catalog_set = set(incomplete_catalog_numbers)

    excel_catalog_map = {}

    for row in range(2, sheet.max_row + 1):
        raw_catalog = sheet.cell(row=row, column=catalog_column).value
        catalog_number = normalize_catalog_number(raw_catalog)

        if catalog_number:
            if catalog_number not in excel_catalog_map:
                excel_catalog_map[catalog_number] = []
            excel_catalog_map[catalog_number].append(row)

    best_item_by_catalog = build_best_item_by_catalog(json_entries)

    matched_entries = 0
    updated_cells = 0
    unmatched_catalogs = []

    for catalog in incomplete_catalog_numbers:
        if catalog not in incomplete_catalog_set:
            continue

        json_item = best_item_by_catalog.get(catalog)
        target_rows = excel_catalog_map.get(catalog, [])

        if not json_item or not target_rows:
            if not json_item:
                unmatched_catalogs.append(catalog)
            continue

        matched_entries += 1
        payload = build_payload(json_item)

        for target_row in target_rows:
            for header_name, new_value in payload.items():
                if not new_value:
                    continue

                target_column = headers.get(header_name)
                if not target_column or target_column < start_column:
                    continue

                cell = sheet.cell(row=target_row, column=target_column)
                if cell.value not in (None, ""):
                    continue

                cell.value = new_value
                updated_cells += 1

    workbook.save(excel_path)
    return matched_entries, updated_cells, len(unmatched_catalogs)


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Populate empty cells in STEELITE.xlsx using goforgreenuk scraped data, "
            "for catalog numbers listed in incomplete_catalog_numbers_v0.8.json, "
            "limited to columns E onward."
        )
    )
    parser.add_argument(
        "--excel",
        default="sources/STEELITE.xlsx",
        help="Path to STEELITE.xlsx",
    )
    parser.add_argument(
        "--json",
        default="steelite/goforgreenuk_v0.3.json",
        help="Path to goforgreenuk JSON file",
    )
    parser.add_argument(
        "--incomplete",
        default="incomplete_catalog_numbers_v0.8.json",
        help="Path to incomplete catalog numbers JSON-like file",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    json_path = Path(args.json)
    incomplete_path = Path(args.incomplete)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if not json_path.exists():
        raise FileNotFoundError(f"JSON file not found: {json_path}")
    if not incomplete_path.exists():
        raise FileNotFoundError(f"Incomplete catalog file not found: {incomplete_path}")

    json_entries = load_json_entries(json_path)
    incomplete_catalog_numbers = load_incomplete_catalog_numbers(incomplete_path)
    matched_entries, updated_cells, unmatched_count = populate_workbook(
        excel_path,
        json_entries,
        incomplete_catalog_numbers,
    )

    print(f"Matched JSON entries: {matched_entries}")
    print(f"Cells populated: {updated_cells}")
    print(f"Incomplete catalog numbers loaded: {len(incomplete_catalog_numbers)}")
    print(f"Incomplete catalog numbers without goforgreenuk match: {unmatched_count}")
    print("Done. Existing data was preserved.")


if __name__ == "__main__":
    main()
