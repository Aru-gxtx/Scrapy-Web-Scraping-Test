from openpyxl import load_workbook
from difflib import SequenceMatcher
import re

def norm_text(v):
    if v is None: return ''
    text = str(v).strip().lower()
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'[^a-z0-9 ]', '', text)
    return text

wb = load_workbook('sources/STEELITE.xlsx', data_only=True)
ws = wb.active
headers = [(ws.cell(1,c).value or '').strip() for c in range(1, min(30, ws.max_column+1))]
cat_col = headers.index('Mfr Catalog No.') + 1
desc_col = headers.index('Item Description') + 1

# Collect all blank-catalog rows
blank_rows = []
for r in range(2, ws.max_row+1):
    cat = ws.cell(r, cat_col).value
    desc = ws.cell(r, desc_col).value
    if (not cat or str(cat).lower() == 'nan') and desc and str(desc).strip():
        blank_rows.append((r, norm_text(desc)))

# Test the 3 products
products = [
    'american almond caramelized almond brittle crunch topping 30',
    'flash furniture fairway black adjustable height office desk',
    'kontos tandoori naan 120 case'
]

for prod_name in products:
    print(f'\nProduct: {prod_name}')
    prod_norm = norm_text(prod_name)
    scores = []
    for row, desc_norm in blank_rows:
        score = SequenceMatcher(None, prod_norm, desc_norm).ratio()
        substring_bonus = 0.2 if any(word in desc_norm.split() for word in prod_norm.split()) else 0
        total = score + substring_bonus
        scores.append((total, row, score, substring_bonus, desc_norm[:50]))
    
    scores.sort(reverse=True)
    print(f'Top 3 matches:')
    for total, row, score, bonus, desc in scores[:3]:
        print(f'  Row {row}: {total:.3f} (base={score:.3f}, bonus={bonus:.2f}) | {desc}')
