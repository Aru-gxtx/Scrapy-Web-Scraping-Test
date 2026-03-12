from pathlib import Path

from openpyxl import load_workbook


INPUT_FILE = Path("sources/STEELITE.xlsx")
OUTPUT_FILE = Path("sources/STEELITE_Updated_filled.xlsx")
START_COL = 5   # Column E
END_COL = 20    # Column T
FILL_VALUE = "n/a"


def is_blank(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    workbook = load_workbook(INPUT_FILE)
    filled_count = 0

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=START_COL,
            max_col=END_COL,
        ):
            for cell in row:
                if is_blank(cell.value):
                    cell.value = FILL_VALUE
                    filled_count += 1

    workbook.save(OUTPUT_FILE)
    print(f"Done. Filled {filled_count} blank cells from E to T.")
    print(f"Saved: {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()
