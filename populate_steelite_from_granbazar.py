import argparse
import json
import re
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import unquote

from openpyxl import load_workbook


def normalize_catalog_number(value):
    if value is None:
        return ""

    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""

    text = unquote(text)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text


def extract_catalog_from_product_name(product_name):
    text = str(product_name or "")
    if not text:
        return ""

    # Example: "... Steelite 1107 0569" -> "11070569"
    match = re.search(r"steelite\s+([A-Z0-9]{4})\s+([A-Z0-9]{4})\b", text, re.IGNORECASE)
    if match:
        return normalize_catalog_number(f"{match.group(1)}{match.group(2)}")

    # Fallback: sometimes already contiguous near "Steelite"
    match = re.search(r"steelite\s+([A-Z0-9]{8,})\b", text, re.IGNORECASE)
    if match:
        return normalize_catalog_number(match.group(1))

    return ""


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^a-z0-9 ]", "", text)
    return text


CYRILLIC_TO_LATIN = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
    "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
    "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "kh", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "shch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
}


WORD_REPLACEMENTS = {
    "фарфор": "porcelain",
    "стекло": "glass",
    "керамика": "ceramic",
    "пластик": "plastic",
    "белый": "white",
    "черный": "black",
    "красный": "red",
    "синий": "blue",
    "зеленый": "green",
    "зелёный": "green",
    "желтый": "yellow",
    "жёлтый": "yellow",
    "серый": "gray",
    "серебристый": "silver",
    "золотой": "gold",
    "бежев": "beige",
    "коричнев": "brown",
    "чашка": "cup",
    "блюдце": "saucer",
    "салатник": "salad bowl",
    "тарелка": "plate",
    "кружка": "mug",
    "миска": "bowl",
    "крышка": "lid",
    "мелкая": "flat",
    "бульонная": "bouillon",
    "столовая": "dinner",
}


PHRASE_REPLACEMENTS = [
    (r"купить в москве, санкт-петербурге и краснодаре с доставкой по выгодной цене", "available with delivery at competitive price"),
    (r"вы можете на сайте поставщика granбазар", "on the GranBazar supplier website"),
    (r"после оформления заказа", "after order placement"),
    (r"наш менеджер свяжется с вами для уточнения времени доставки и деталей оплаты", "our manager will contact you to confirm delivery time and payment details"),
    (r"мы предлагаем продажу товара с гарантией от продавца и производителя", "this product is provided with seller and manufacturer warranty"),
    (r"доставка", "delivery"),
]


def transliterate_cyrillic(text):
    result = []
    for ch in str(text or ""):
        lower = ch.lower()
        if lower in CYRILLIC_TO_LATIN:
            latin = CYRILLIC_TO_LATIN[lower]
            result.append(latin.capitalize() if ch.isupper() and latin else latin)
        else:
            result.append(ch)
    return "".join(result)


def translate_russian_text(text):
    value = str(text or "").strip()
    if not value:
        return ""

    # Normalize common units first.
    value = re.sub(r"\bмм\b", "mm", value, flags=re.IGNORECASE)
    value = re.sub(r"\bмл\b", "ml", value, flags=re.IGNORECASE)
    value = re.sub(r"(?<=\d)\s*л\b", " l", value, flags=re.IGNORECASE)

    lowered = value.lower()
    for pattern, replacement in PHRASE_REPLACEMENTS:
        lowered = re.sub(pattern, replacement, lowered, flags=re.IGNORECASE)

    # Replace common words and stems.
    for src, dst in WORD_REPLACEMENTS.items():
        lowered = re.sub(rf"\b{re.escape(src)}[\w.]*", dst, lowered, flags=re.IGNORECASE)

    # Transliterate any remaining Cyrillic text.
    translated = transliterate_cyrillic(lowered)
    translated = re.sub(r"\s+", " ", translated).strip(" ;,")
    return translated


def translate_simple_value(value):
    text = str(value or "").strip()
    if not text:
        return ""
    return translate_russian_text(text)


def build_payload(item):
    product_name = translate_russian_text((item.get("product_name") or "").strip())
    overview = translate_russian_text((item.get("overview") or "").strip())

    if not overview:
        overview = product_name

    payload = {
        "Image Link": (item.get("image_link") or "").strip(),
        "Overview": overview,
        "Length": (item.get("length") or "").strip(),
        "Width": (item.get("width") or "").strip(),
        "Height": (item.get("height") or "").strip(),
        "Capacity": (item.get("capacity") or item.get("volume") or "").strip(),
        "Features": "",
        "Edge Style": "",
        "Volume": (item.get("volume") or item.get("capacity") or "").strip(),
        "Diameter": (item.get("diameter") or "").strip(),
        "Color": translate_simple_value(item.get("color")),
        "Shape": translate_simple_value(item.get("shape")),
        "Material": translate_simple_value(item.get("material")),
        "Pattern": translate_simple_value(item.get("pattern")),
        "EAN Code": (item.get("ean_code") or "").strip(),
        "Barcode": (item.get("barcode") or "").strip(),
        # Optional columns: written only when present in the workbook headers.
        "Price": (item.get("price") or "").strip(),
        "Product URL": (item.get("product_url") or "").strip(),
        "Page Number": str(item.get("page_number") or "").strip(),
    }

    return payload


def load_json_entries(json_path):
    with open(json_path, "r", encoding="utf-8") as file:
        return json.load(file)


def populate_workbook(excel_path, json_entries):
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    headers = {
        (sheet.cell(row=1, column=column).value or "").strip(): column
        for column in range(1, sheet.max_column + 1)
    }

    catalog_column = headers.get("Mfr Catalog No.")
    if not catalog_column:
        raise ValueError("Column 'Mfr Catalog No.' was not found in the workbook.")

    item_description_column = headers.get("Item Description")
    start_column = 5

    excel_catalog_map = {}
    blank_catalog_rows = []

    for row in range(2, sheet.max_row + 1):
        raw_catalog = sheet.cell(row=row, column=catalog_column).value
        catalog_number = normalize_catalog_number(raw_catalog)

        if catalog_number:
            if catalog_number not in excel_catalog_map:
                excel_catalog_map[catalog_number] = []
            excel_catalog_map[catalog_number].append(row)
            continue

        if item_description_column:
            raw_description = sheet.cell(row=row, column=item_description_column).value
            normalized_description = normalize_text(raw_description)
            if normalized_description:
                blank_catalog_rows.append((row, normalized_description))

    matched_entries = 0
    updated_cells = 0
    used_excel_rows = set()

    for json_item in json_entries:
        if json_item.get("found") is False:
            continue

        title_catalog = extract_catalog_from_product_name(json_item.get("product_name"))
        json_catalog = normalize_catalog_number(
            json_item.get("catalog_number")
            or json_item.get("searched_catalog_number")
            or title_catalog
        )

        # Last-resort fallbacks when no catalog-like code is found in title.
        if not json_catalog:
            json_catalog = normalize_catalog_number(
                json_item.get("ean_code")
                or json_item.get("product_id")
            )
        target_row = None

        if json_catalog:
            matching_excel_rows = excel_catalog_map.get(json_catalog, [])

            for row in matching_excel_rows:
                if row not in used_excel_rows:
                    target_row = row
                    used_excel_rows.add(row)
                    break

            if not target_row and matching_excel_rows:
                target_row = matching_excel_rows[0]
        else:
            product_name = normalize_text(json_item.get("product_name"))
            if product_name and blank_catalog_rows:
                best_row = None
                best_score = 0.0

                for row, description in blank_catalog_rows:
                    if row in used_excel_rows:
                        continue

                    score = SequenceMatcher(None, product_name, description).ratio()
                    if product_name in description or description in product_name:
                        score += 0.2

                    if score > best_score:
                        best_score = score
                        best_row = row

                if best_row and best_score >= 0.55:
                    target_row = best_row
                    used_excel_rows.add(best_row)

        if not target_row:
            continue

        matched_entries += 1
        payload = build_payload(json_item)

        for header_name, new_value in payload.items():
            if not new_value:
                continue

            target_column = headers.get(header_name)
            if not target_column or target_column < start_column:
                continue

            cell = sheet.cell(row=target_row, column=target_column)
            if cell.value not in (None, ""):
                continue

            cell.value = new_value
            updated_cells += 1

    workbook.save(excel_path)
    return matched_entries, updated_cells


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Populate empty cells in STEELITE.xlsx using granbazar scraped data, "
            "matched by 'Mfr Catalog No.' and limited to columns E onward."
        )
    )
    parser.add_argument(
        "--excel",
        default="sources/STEELITE.xlsx",
        help="Path to STEELITE.xlsx",
    )
    parser.add_argument(
        "--json",
        default="steelite/granbazar.json",
        help="Path to granbazar JSON file",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    json_path = Path(args.json)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if not json_path.exists():
        raise FileNotFoundError(f"JSON file not found: {json_path}")

    json_entries = load_json_entries(json_path)
    matched_entries, updated_cells = populate_workbook(excel_path, json_entries)

    print(f"Matched JSON entries: {matched_entries}")
    print(f"Cells populated: {updated_cells}")
    print("Done. Existing data was preserved.")


if __name__ == "__main__":
    main()